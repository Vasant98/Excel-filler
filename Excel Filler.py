# -*- coding: utf-8 -*-
"""
Created on Thu Apr 21 14:16:05 2022

@author: WangSheng
"""


import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from pdfrw  import PDFWriter

wb = load_workbook(r'C:\Users\Phang In Yee\Downloads\OneDrive_1_4-25-2022\F-826-01-004_Rev 0_Certificate of Conformance Template.xlsx')

input_data = os.path.normpath(os.path.expanduser
                             (r"C:\Users\Phang In Yee\Downloads\OneDrive_1_4-25-2022\New_V2 Product Lot & SN release date.csv"))


df = pd.read_csv(input_data)

input_df = df[['SN', 'DATE']]

for i in input_df.index:
    SN = df['SN'][i]
    DATE = df['DATE'][i]
    
    ws = wb.active
    ws['A18'].value = SN
    ws['F41'].value = DATE
    
    wb.save('F-826-01-004_Rev 0_COC_' + SN + '.xlsx')
    
    pw = PDFWriter('F-826-01-004_Rev 0_COC_' + SN + '.pdf')
    # pw.setFont('Courier', 12)
    # pw.setHeader('XLSXtoPDF.py - convert XLSX data to PDF')
    # pw.setFooter('Generated using openpyxl and xtopdf')

    ws_range = wb.iter_rows('A1:F13')
    for row in ws_range:
        s = ''
        for cell in row:
            if cell.value is None:
                s += ' ' * 11
            else:
                s += str(cell.value).rjust(10) + ' '
        pw.writeLine(s)
    pw.savePage()
    pw.close()